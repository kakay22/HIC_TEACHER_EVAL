import django
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Section, Question, Teacher, Course, Subject, TeacherEvaluation, EvaluationItem, Student
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from .forms import SectionForm, QuestionForm
from django.core.paginator import Paginator
from .forms import CourseForm
from django.db.models import Count
from django.http import JsonResponse

from django.db.models import Q
from django.http import HttpResponse
import csv
from reportlab.platypus import SimpleDocTemplate, Table

from django.contrib.staticfiles import finders

def home(request):
    sections = Section.objects.prefetch_related('questions').all()
    teachers = Teacher.objects.all()
    courses = Course.objects.all()
    subjects = Subject.objects.all()
    years = ["2023-2024", "2024-2025", "2025-2026", "2026-2027", "2027-2028"]  # or generate dynamically
    year_levels = ["1st Year", "2nd Year", "3rd Year", "4th Year", "5th Year"]

    if request.method == "POST":
        # Get form data
        student_id = request.POST.get("student_id")
        student_name = request.POST.get("student_name")  # optional
        teacher_id = request.POST.get("teacher")
        semester = request.POST.get("semester")
        academic_year = request.POST.get("academic_year")
        subject_id = request.POST.get("subject")
        course_id = request.POST.get("course")
        year_level = request.POST.get("year_level")

        # Additional comments
        commendable_features = request.POST.get("commendable_features", "")
        suggestions_improvement = request.POST.get("suggestions_improvement", "")

        # Get related objects
        teacher = Teacher.objects.get(id=teacher_id)
        subject = Subject.objects.get(id=subject_id)
        course = Course.objects.get(id=course_id)

        # Check if this student already submitted for this teacher/subject/semester/year
        existing = TeacherEvaluation.objects.filter(
            student_id=student_id,
            teacher=teacher,
            subject=subject,
            semester=semester,
            academic_year=academic_year
        ).exists()

        if existing:
            messages.error(request, "You have already submitted an evaluation for this teacher, subject, and academic year.")
            return redirect("home")

        # Create evaluation
        evaluation = TeacherEvaluation.objects.create(
            student_id=student_id,
            student_name=student_name,
            teacher=teacher,
            semester=semester,
            academic_year=academic_year,
            subject=subject,
            course=course,
            year_level=year_level,
            commendable_features=commendable_features,
            suggestions_improvement=suggestions_improvement
        )

        # Save all ratings
        for section in sections:
            prefix = section.prefix
            for question in section.questions.all():
                field_name = f"q_{prefix}_{question.order}"
                rating = request.POST.get(field_name)
                if rating:
                    EvaluationItem.objects.create(
                        evaluation=evaluation,
                        question=question,
                        rating=rating
                    )

        messages.success(request, "Your evaluation has been submitted successfully!")
        return redirect("thank_you")  # Redirect to thank you page

    context = {
        "sections": sections,
        "teachers": teachers,
        "courses": courses,
        "subjects": subjects,
        "years": years,
        "year_levels": year_levels,
    }
    return render(request, 'evaluation/home.html', context)

def validate_evaluation(request):
    # Only allow AJAX requests
    if request.headers.get("X-Requested-With") != "XMLHttpRequest":
        return JsonResponse({"student_exists": False, "duplicate": False})

    student_id = request.GET.get("student_id")
    teacher = request.GET.get("teacher")
    subject = request.GET.get("subject")
    semester = request.GET.get("semester")
    academic_year = request.GET.get("academic_year")

    # Check if student ID exists
    student_exists = Student.objects.filter(student_id=student_id).exists() if student_id else False

    # Check for duplicate only if all fields are filled
    duplicate = False
    if all([student_id, teacher, subject, semester, academic_year]):
        duplicate = TeacherEvaluation.objects.filter(
            student_id=student_id,
            teacher_id=teacher,
            subject_id=subject,
            semester=semester,
            academic_year=academic_year
        ).exists()

    return JsonResponse({
        "student_exists": student_exists,
        "duplicate": duplicate
    })



def thank_you(request):
    return render(request, 'evaluation/thankyou.html')

from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect

def admin_login(request):

    username = ""
    password = ""
    
    if request.user.is_authenticated:
        return redirect('dashboard_home')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember = request.POST.get('remember')  # ✅ get checkbox

        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_superuser:
            login(request, user)

            # ✅ REMEMBER ME LOGIC
            if not remember:
                request.session.set_expiry(0)  # expires on browser close
            else:
                request.session.set_expiry(1209600)  # 2 weeks

            messages.success(request, "Login successful!")
            return redirect('dashboard_home')

        else:
            messages.error(request, "Invalid credentials or not authorized.")

    return render(request, 'evaluation/admin_login.html', {
        'username': username,
        'password': password
    })


# Admin logout
def admin_logout(request):
    logout(request)
    messages.success(request, "Logged out successfully.")
    return redirect('admin_login')

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg, Q, FloatField
from django.db.models.functions import Cast
from django.utils import timezone
from datetime import timedelta

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Count
from django.utils import timezone
from datetime import timedelta
from .models import TeacherEvaluation, Teacher, EvaluationItem

@login_required(login_url='login')
def dashboard_home(request):
    semester = request.GET.get('semester')
    academic_year = request.GET.get('academic_year')
    teacher_id = request.GET.get('teacher')

    evaluations_qs = TeacherEvaluation.objects.select_related(
        'teacher', 'subject', 'course'
    )

    if semester:
        evaluations_qs = evaluations_qs.filter(semester=semester)
    if academic_year:
        evaluations_qs = evaluations_qs.filter(academic_year=academic_year)
    if teacher_id:
        evaluations_qs = evaluations_qs.filter(teacher_id=teacher_id)

    # ===============================
    # BASIC SUMMARY COUNTS
    # ===============================
    total_teachers = Teacher.objects.count()
    total_evaluations = evaluations_qs.count()
    total_students = evaluations_qs.values('student_id').distinct().count()

    # ===============================
    # AVERAGE RATING
    # ===============================
    avg_rating_value = evaluations_qs.aggregate(avg=Avg('overall_rating'))['avg']
    avg_rating_value = round(avg_rating_value, 2) if avg_rating_value is not None else 0

    # ===============================
    # RATING DISTRIBUTION
    # ===============================
    rating_counts_qs = (
        EvaluationItem.objects
        .filter(evaluation__in=evaluations_qs)
        .exclude(rating='N')
        .values('rating')
        .annotate(count=Count('id'))
    )

    rating_data = {str(i): 0 for i in range(1, 6)}
    for r in rating_counts_qs:
        rating_data[str(r['rating'])] = r['count']

    # ===============================
    # RECENT EVALUATIONS
    # ===============================
    recent_evaluations = evaluations_qs.order_by('-submitted_at')[:5]

    # ===============================
    # COMPLETION & RESPONSE RATE
    # ===============================
    completion_rate = 100 if total_students else 0
    response_rate = round(total_evaluations / total_students, 2) if total_students else 0

    # ===============================
    # WEEKLY GROWTH RATE
    # ===============================
    today = timezone.now()
    last_7_days = today - timedelta(days=7)
    previous_7_days = today - timedelta(days=14)

    recent_count = TeacherEvaluation.objects.filter(
        submitted_at__gte=last_7_days
    ).count()

    previous_count = TeacherEvaluation.objects.filter(
        submitted_at__gte=previous_7_days,
        submitted_at__lt=last_7_days
    ).count()

    growth_rate = round(
        ((recent_count - previous_count) / previous_count) * 100, 1
    ) if previous_count else 0

    # ===============================
    # TOP TEACHER
    # ===============================
    top_teacher_data = (
        evaluations_qs
        .values('teacher__id', 'teacher__first_name', 'teacher__last_name')
        .annotate(avg=Avg('overall_rating'), total=Count('id'))
        .order_by('-avg')
        .first()
    )

    top_teacher = None
    top_teacher_avg = 0
    if top_teacher_data and top_teacher_data['avg'] is not None:
        top_teacher = top_teacher_data
        top_teacher_avg = round(top_teacher_data['avg'], 2)

    # ===============================
    # QUICK STATS
    # ===============================
    start_of_month = today.replace(day=1)
    monthly_evaluations = evaluations_qs.filter(submitted_at__gte=start_of_month).count()
    active_teachers = evaluations_qs.values('teacher').distinct().count()
    highest_rating = 5.0

    # ===============================
    # STATUS COUNTS
    # ===============================
    completed_count = total_evaluations
    pending_count = 0   # ❗ requires expected submissions to calculate
    inprogress_count = 0

    # ===============================
    # CONTEXT
    # ===============================
    context = {
        'total_teachers': total_teachers,
        'total_evaluations': total_evaluations,
        'total_students': total_students,

        'avg_rating': avg_rating_value,
        'rating_data': rating_data,
        'recent_evaluations': recent_evaluations,

        'completion_rate': completion_rate,
        'growth_rate': growth_rate,
        'top_teacher': top_teacher,
        'top_teacher_avg': top_teacher_avg,

        'completed_count': completed_count,
        'pending_count': pending_count,
        'inprogress_count': inprogress_count,

        'monthly_evaluations': monthly_evaluations,
        'active_teachers': active_teachers,
        'highest_rating': highest_rating,
        'response_rate': response_rate,

        'selected_semester': semester or '',
        'selected_academic_year': academic_year or '',
        'selected_teacher': teacher_id or '',
        'teachers': Teacher.objects.all(),
    }

    return render(request, 'admins/dashboard_home.html', context)

from django.db.models import Q

@login_required(login_url='login')
def teachers_list(request):
    query = request.GET.get('q')

    teachers = Teacher.objects.all()  # ✅ NO select_related here

    if query:
        teachers = teachers.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )

    context = {
        'teachers': teachers,
    }
    return render(request, 'admins/teachers.html', context)

from django.shortcuts import get_object_or_404, render
from django.db.models import Avg
from .models import Teacher, EvaluationItem 

def teacher_detail(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    evaluations = teacher.evaluations.prefetch_related('items').all()

    # Calculate overall average rating
    # Convert string ratings '1'..'5' to float
    all_ratings = []
    for eval in evaluations:
        for item in eval.items.all():
            if item.rating in ['1','2','3','4','5']:
                all_ratings.append(int(item.rating))
    avg_rating = round(sum(all_ratings)/len(all_ratings),2) if all_ratings else 0

    context = {
        'teacher': teacher,
        'teacher_evaluations': evaluations,
        'avg_rating': avg_rating
    }
    return render(request, 'admins/teacher_details.html', context)

def courses_list(request):
    # Optional search
    query = request.GET.get('q', '')
    if query:
        courses = Course.objects.filter(name__icontains=query).order_by('name')
    else:
        courses = Course.objects.all().order_by('name')

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(courses, 20)  # 20 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'courses': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
        'query': query,
    }
    return render(request, 'admins/courses_list.html', context)


from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
import csv
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table
from .models import TeacherEvaluation, Teacher, Subject
from django.db.models import F


from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .models import TeacherEvaluation, Teacher, Subject

@login_required(login_url='admin_login')
def evaluations_list(request):
    # ===================== QUERYSET =====================
    qs = TeacherEvaluation.objects.select_related('teacher', 'subject', 'course').prefetch_related('items')

    # ===================== FILTERS =====================
    teacher_id = request.GET.get('teacher')
    subject_id = request.GET.get('subject')
    semester = request.GET.get('semester')
    search = request.GET.get('search')
    sort = request.GET.get('sort')
    export = request.GET.get('export')

    if teacher_id:
        qs = qs.filter(teacher_id=teacher_id)
    if subject_id:
        qs = qs.filter(subject_id=subject_id)
    if semester:
        qs = qs.filter(semester=semester)

    # ===================== SEARCH =====================
    if search:
        qs = qs.filter(
            Q(teacher__first_name__icontains=search) |
            Q(teacher__last_name__icontains=search) |
            Q(subject__name__icontains=search)
        )

    # ===================== UPDATE OVERALL_RATING (BULK) =====================
    updates = []
    for e in qs:
        # Assign overall_rating from computed value
        if hasattr(e, 'overall_computed_rating') and e.overall_computed_rating is not None:
            e.overall_rating = e.overall_computed_rating

        # Compute stars for template
        rating = e.overall_rating or 0
        filled = int(rating)
        half = 1 if (rating - filled) >= 0.5 else 0
        empty = 5 - filled - half

        e.filled_stars_list = range(filled)
        e.half_star_list = range(half)
        e.empty_stars_list = range(empty)

        updates.append(e)

    # Bulk update overall_rating in DB
    if updates:
        TeacherEvaluation.objects.bulk_update(updates, ['overall_rating'])

    # ===================== SORT =====================
    if sort == "date_desc":
        qs = qs.order_by('-submitted_at')
    elif sort == "date_asc":
        qs = qs.order_by('submitted_at')
    elif sort == "rating_desc":
        qs = sorted(qs, key=lambda x: (getattr(x, 'overall_computed_rating', 0) or 0), reverse=True)
    elif sort == "rating_asc":
        qs = sorted(qs, key=lambda x: (getattr(x, 'overall_computed_rating', 0) or 0))

    # ===================== EXPORT =====================
    if export in ["excel", "pdf"]:
        for e in qs:
            ratings = []
            for item in e.items.all():
                value = str(getattr(item, 'rating', '')).strip().upper()
                if value == "N":
                    continue
                try:
                    num = float(value)
                    if 1 <= num <= 5:
                        ratings.append(num)
                except:
                    continue
            e.computed_rating = round(sum(ratings)/len(ratings), 2) if ratings else None

        if export == "excel":
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Evaluations"

            headers = ['Teacher', 'Subject', 'Overall Rating', 'Semester', 'A.Y.', 'Submitted']
            ws.append(headers)

            # Header styling
            for col in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

            # Data rows
            for e in qs:
                ws.append([
                    f"{e.teacher.first_name} {e.teacher.last_name}" if e.teacher else "N/A",
                    e.subject.name if e.subject else "N/A",
                    e.computed_rating or "N/A",
                    e.semester,
                    e.academic_year,
                    e.submitted_at.strftime('%Y-%m-%d')
                ])

            # Auto-width
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

            ws.freeze_panes = "A2"

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Evaluations.xlsx'
            wb.save(response)
            return response

        elif export == "pdf":
            from io import BytesIO
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            elements.append(Paragraph("Evaluations Report", styles['Title']))
            elements.append(Spacer(1,12))

            data = [['Teacher', 'Subject', 'Overall Rating', 'Semester', 'A.Y.', 'Submitted']]
            for e in qs:
                data.append([
                    f"{e.teacher.first_name} {e.teacher.last_name}" if e.teacher else "N/A",
                    e.subject.name if e.subject else "N/A",
                    e.computed_rating or "N/A",
                    e.semester,
                    e.academic_year,
                    e.submitted_at.strftime('%Y-%m-%d')
                ])

            table = Table(data, colWidths=[2.5*inch,2*inch,1*inch,1*inch,1*inch,1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F46E5")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN',(2,1),(-1,-1),'CENTER'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.gray),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ]))
            elements.append(table)
            doc.build(elements)

            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=Evaluations.pdf'
            return response

    # ===================== PAGINATION =====================
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    # ===================== CONTEXT =====================
    context = {
        'page_obj': page_obj,
        'teachers': Teacher.objects.all(),
        'subjects': Subject.objects.all(),
        'selected_teacher': teacher_id or '',
        'selected_subject': subject_id or '',
        'selected_semester': semester or '',
        'search': search or '',
        'sort': sort or '',
    }

    return render(request, 'admins/evaluations.html', context)


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import TeacherEvaluation
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import letter, inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from .models import TeacherEvaluation

@login_required(login_url='admin_login')
def evaluation_detail(request, evaluation_id):
    from django.contrib.staticfiles import finders
    from io import BytesIO
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.lib import colors

    evaluation = get_object_or_404(
        TeacherEvaluation.objects.prefetch_related('items__question__section'),
        id=evaluation_id
    )

    # ------------------- BUILD SECTIONS DICT -------------------
    sections_dict = {}
    for item in evaluation.items.all():
        section = item.question.section
        if section.prefix not in sections_dict:
            sections_dict[section.prefix] = {
                'title': section.title,
                'items': [],
                'total': 0
            }
        sections_dict[section.prefix]['items'].append(item)
        try:
            rating_num = int(item.rating)
            sections_dict[section.prefix]['total'] += rating_num
        except ValueError:
            pass

    # ------------------- EXPORT LOGIC -------------------
    export_type = request.GET.get('export', None)
    teacher_name = f"{evaluation.teacher.first_name}_{evaluation.teacher.last_name}".replace(" ", "_")

    # ------------------- EXCEL EXPORT -------------------
    if export_type == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluation"

        # Header
        ws.append([f"Teacher Evaluation Report"])
        ws.append([f"Teacher: {evaluation.teacher.first_name} {evaluation.teacher.last_name}"])
        ws.append([f"Semester: {evaluation.semester} | Academic Year: {evaluation.academic_year}"])
        ws.append([f"Submitted At: {evaluation.submitted_at.strftime('%B %d, %Y %I:%M %p')}"])
        ws.append([])

        # Column titles
        ws.append(["Section", "Question", "Rating"])
        for col in range(1, 4):
            ws.cell(row=6, column=col).font = Font(bold=True)
            ws.cell(row=6, column=col).alignment = Alignment(horizontal='center')

        row_num = 7
        for section in sections_dict.values():
            ws.cell(row=row_num, column=1, value=section['title']).font = Font(bold=True)
            row_num += 1
            for idx, item in enumerate(section['items']):
                ws.append(["", f"{item.question.order}. {item.question.text}", item.rating])
                # Alternate row fill
                if idx % 2 == 0:
                    for col in range(1, 4):
                        ws.cell(row=row_num, column=col).fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                row_num += 1
            # Section total
            ws.append(["", "Section Total", section['total']])
            for col in range(2, 4):
                ws.cell(row=row_num, column=col).font = Font(bold=True)
                ws.cell(row=row_num, column=col).fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            row_num += 2

        # Additional Comments
        if evaluation.commendable_features:
            ws.append(["Additional Comments", "Commendable Features", ""])
            ws.append(["", evaluation.commendable_features, ""])
            row_num += 2
        if evaluation.suggestions_improvement:
            ws.append(["Additional Comments", "Suggestions for Improvement", ""])
            ws.append(["", evaluation.suggestions_improvement, ""])
            row_num += 2

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 5

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={teacher_name}_evaluation.xlsx'
        wb.save(response)
        return response

    # ------------------- PDF EXPORT -------------------
    elif export_type == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        # ------------------- HEADER: LOGO + TITLE CENTERED -------------------
        logo_path = finders.find('HIC_LOGO_TRANSPARENT.png')
        logo_img = None
        if logo_path:
            # slightly smaller logo
            logo_img = Image(logo_path, width=1.0*inch, height=1.0*inch)

        header_data = []
        if logo_img:
            header_data.append([logo_img, Paragraph(
                "<b>Teacher Evaluation Report</b><br/><font size=12>Academic Evaluation Record</font>",
                styles['Title']
            )])
            col_widths = [1.0*inch, 4.5*inch]
        else:
            header_data.append([Paragraph(
                "<b>Teacher Evaluation Report</b><br/><font size=12>Academic Evaluation Record</font>",
                styles['Title']
            )])
            col_widths = [6*inch]

        header_table = Table(header_data, colWidths=col_widths)
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6)
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 18))  # balanced spacing after header

        # ------------------- BASIC INFO -------------------
        info_data = [
            ["Subject", evaluation.subject.name if evaluation.subject else "N/A"],
            ["Teacher", f"{evaluation.teacher.first_name} {evaluation.teacher.last_name}"],
            ["Semester", evaluation.semester],
            ["Academic Year", evaluation.academic_year],
            ["Year Level", evaluation.year_level],
            ["Student ID", evaluation.student_id]
        ]
        if evaluation.student_name:
            info_data.append(["Student Name", evaluation.student_name])
        info_data.append(["Date Submitted", evaluation.submitted_at.strftime('%B %d, %Y %I:%M %p')])

        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEBELOW', (1,0), (1,-1), 0.5, colors.gray)
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 12))

        # ------------------- SECTIONS -------------------
        for prefix, section in sections_dict.items():
            elements.append(Paragraph(f"{prefix}. {section['title']}", styles['Heading3']))
            section_data = [["Question", "Rating"]]
            for idx, item in enumerate(section['items']):
                section_data.append([f"{item.question.order}. {item.question.text}", str(item.rating)])
            section_data.append(["Total Score", str(section['total'])])

            section_table = Table(section_data, colWidths=[4*inch, 2*inch])
            style_list = [
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F46E5")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN',(1,1),(-1,-1),'CENTER'),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.gray),
                ('BOX', (0,0), (-1,-1), 1, colors.black),
            ]
            for i in range(1, len(section_data)-1):
                if i % 2 == 0:
                    style_list.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor("#f3f4f6")))
            style_list.append(('BACKGROUND', (0,len(section_data)-1), (-1,len(section_data)-1), colors.HexColor("#e5e7eb")))
            style_list.append(('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'))
            style_list.append(('FONTNAME', (0,len(section_data)-1), (-1,len(section_data)-1), 'Helvetica-Bold'))
            section_table.setStyle(TableStyle(style_list))
            elements.append(section_table)
            elements.append(Spacer(1, 12))

        # ------------------- ADDITIONAL COMMENTS -------------------
        if evaluation.commendable_features or evaluation.suggestions_improvement:
            elements.append(Paragraph("Additional Comments", styles['Heading2']))
            elements.append(Spacer(1, 6))
            if evaluation.commendable_features:
                elements.append(Paragraph("A. Commendable Features", styles['Heading4']))
                elements.append(Paragraph(evaluation.commendable_features, styles['Normal']))
                elements.append(Spacer(1, 6))
            if evaluation.suggestions_improvement:
                elements.append(Paragraph("B. Suggestions for Improvement", styles['Heading4']))
                elements.append(Paragraph(evaluation.suggestions_improvement, styles['Normal']))
                elements.append(Spacer(1, 6))

        # ------------------- BUILD PDF -------------------
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename={teacher_name}_evaluation.pdf'
        return response

    # ------------------- NORMAL PAGE VIEW -------------------
    context = {
        'evaluation': evaluation,
        'sections': sections_dict
    }
    return render(request, 'admins/evaluation_detail.html', context)

# ------------------ SECTIONS ------------------
@login_required(login_url='admin_login')
def sections_list(request):
    sections = Section.objects.prefetch_related('questions').all()
    return render(request, 'admins/sections.html', {'sections': sections})

@login_required(login_url='admin_login')
def section_add(request):
    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section created successfully!')
        else:
            messages.error(request, 'Error creating section.')
    return redirect('sections_list')  # Always redirect back to sections page

@login_required(login_url='admin_login')
def section_edit(request, pk):
    section = get_object_or_404(Section, pk=pk)
    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section updated successfully!')
        else:
            messages.error(request, 'Error updating section.')
    return redirect('sections_list')

@login_required(login_url='admin_login')
def section_delete(request, pk):
    section = get_object_or_404(Section, pk=pk)
    if request.method == 'POST':
        section.delete()
        messages.success(request, 'Section deleted successfully!')
    return redirect('sections_list')

# ------------------ QUESTIONS ------------------
@login_required(login_url='admin_login')
def question_add(request, section_id):
    section = get_object_or_404(Section, pk=section_id)
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.section = section  # assign the section manually
            question.save()
            messages.success(request, "Question added successfully!")
        else:
            # If form errors, show them for debugging
            messages.error(request, f"Error adding question: {form.errors}")
    return redirect('sections_list')


@login_required(login_url='admin_login')
def question_edit(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    if request.method == 'POST':
        new_text = request.POST.get('text', '').strip()
        if new_text:
            question.text = new_text
            question.save()
            messages.success(request, 'Question updated successfully.')
        else:
            messages.error(request, 'Question text cannot be empty.')
    return redirect('sections_list')  # or your URL name for this page


@login_required(login_url='admin_login')
def question_delete(request, pk):
    question = get_object_or_404(Question, pk=pk)
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Question deleted successfully!')
    return redirect('sections_list')


# ------------------- TEACHERS -------------------

@login_required(login_url='admin_login')
def teachers_list(request):
    query = request.GET.get('q')
    teachers = Teacher.objects.all()  # no select_related since no FK fields

    if query:
        teachers = teachers.filter(
            first_name__icontains=query
        ) | teachers.filter(
            last_name__icontains=query
        )

    # Pass subjects and courses for the Add/Edit modals (if you need them for evaluation)
    subjects = Subject.objects.all()
    courses = Course.objects.all()

    context = {
        'teachers': teachers,
        'subjects': subjects,
        'courses': courses,
    }
    return render(request, 'admins/teachers.html', context)


@login_required(login_url='admin_login')
def teacher_add(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")

        # Create the teacher
        Teacher.objects.create(
            first_name=first_name,
            last_name=last_name
        )
        messages.success(request, "Teacher added successfully!")
    return redirect('teachers_list')


@login_required(login_url='admin_login')
def teacher_edit(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)

    if request.method == "POST":
        teacher.first_name = request.POST.get("first_name")
        teacher.last_name = request.POST.get("last_name")

        # Save changes
        teacher.save()
        messages.success(request, "Teacher updated successfully!")
    return redirect('teachers_list')

@login_required(login_url='admin_login')
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == "POST":
        teacher.delete()
        messages.success(request, "Teacher deleted successfully!")
    return redirect('teachers_list')

# ------------------- COURSES -------------------
def courses_list(request):
    query = request.GET.get('q', '')
    courses_qs = Course.objects.all()
    if query:
        courses_qs = courses_qs.filter(name__icontains=query)

    paginator = Paginator(courses_qs.order_by('name'), 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'admins/courses_list.html', {
        'courses': page_obj.object_list,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'query': query,
    })


# ADD Course
def course_add(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Course added successfully!")
        else:
            errors = [e for field_errors in form.errors.values() for e in field_errors]
            # Join them into a single string
            messages.error(request, f"❌ {' '.join(errors)}")
    return redirect('courses_list')


# EDIT Course
def course_edit(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Course updated successfully!")
        else:
            errors = [e for field_errors in form.errors.values() for e in field_errors]
            # Join them into a single string
            messages.error(request, f"❌ {' '.join(errors)}")
            
    return redirect('courses_list')


# DELETE Course
def course_delete(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.delete()
        messages.success(request, "✅ Course deleted successfully!")
    return redirect('courses_list')


# API: Teacher list in JSON for Select2
def teacher_list_json(request):
    """
    Return a JSON list of teachers for Select2 dropdown.
    Supports search via ?q=
    """
    query = request.GET.get('q', '')  # search term
    teachers = Teacher.objects.all()

    if query:
        teachers = teachers.filter(
            first_name__icontains=query
        ) | teachers.filter(
            last_name__icontains=query
        )

    # Limit results to avoid sending too many at once
    teachers = teachers[:50]

    data = [
        {"id": teacher.id, "text": f"{teacher.first_name} {teacher.last_name}"}
        for teacher in teachers
    ]
    return JsonResponse(data, safe=False)


# ----------------- SUBJECT VIEWS -----------------
def subjects_list(request):
    query = request.GET.get('q', '')
    subjects = Subject.objects.all()
    if query:
        subjects = subjects.filter(name__icontains=query)
    return render(request, 'admins/subject_list.html', {'subjects': subjects})

def subject_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            Subject.objects.create(name=name)
            messages.success(request, "Subject added successfully!")
        return redirect('subjects_list')

def subject_edit(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            subject.name = name
            subject.save()
            messages.success(request, "Subject updated successfully!")
        return redirect('subjects_list')

def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        messages.success(request, "Subject deleted successfully!")
        return redirect('subjects_list')

